"""
多云对比和报告生成工具集

提供服务映射、区域映射、价格对比、Excel报告生成等功能。

核心功能：
- 服务映射：建立AWS和Azure服务的对应关系
- 区域映射：将地理位置映射到AWS和Azure区域代码
- 价格对比：对比AWS和Azure的价格差异
- 报告生成：生成包含三个Sheet的Excel报告

作者: Nexus-AI Build Workflow
日期: 2026-01-14
"""

import json
import logging
import os
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from strands import tool

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# 服务映射表
SERVICE_MAPPING = {
    "ec2": {
        "aws_service": "EC2",
        "azure_service": "Virtual Machines",
        "description": "虚拟机实例"
    },
    "ebs": {
        "aws_service": "EBS",
        "azure_service": "Managed Disks",
        "description": "块存储"
    },
    "s3": {
        "aws_service": "S3",
        "azure_service": "Blob Storage",
        "description": "对象存储"
    },
    "rds": {
        "aws_service": "RDS",
        "azure_service": "Azure SQL Database",
        "description": "关系数据库"
    },
    "elasticache": {
        "aws_service": "ElastiCache",
        "azure_service": "Azure Cache for Redis",
        "description": "缓存服务"
    },
    "opensearch": {
        "aws_service": "OpenSearch",
        "azure_service": "Azure Cognitive Search",
        "description": "搜索服务"
    },
    "elb": {
        "aws_service": "ELB",
        "azure_service": "Application Gateway/Load Balancer",
        "description": "负载均衡器"
    },
    "network": {
        "aws_service": "Network",
        "azure_service": "Bandwidth",
        "description": "网络流量"
    }
}


# 区域映射表
REGION_MAPPING = {
    "北美": {
        "美国东部": {"aws": "us-east-1", "azure": "eastus"},
        "美国东部2": {"aws": "us-east-2", "azure": "eastus2"},
        "美国西部": {"aws": "us-west-1", "azure": "westus"},
        "美国西部2": {"aws": "us-west-2", "azure": "westus2"},
        "加拿大中部": {"aws": "ca-central-1", "azure": "canadacentral"}
    },
    "欧洲": {
        "爱尔兰": {"aws": "eu-west-1", "azure": "northeurope"},
        "伦敦": {"aws": "eu-west-2", "azure": "uksouth"},
        "巴黎": {"aws": "eu-west-3", "azure": "francecentral"},
        "法兰克福": {"aws": "eu-central-1", "azure": "germanywestcentral"},
        "斯德哥尔摩": {"aws": "eu-north-1", "azure": "swedencentral"}
    },
    "亚太": {
        "新加坡": {"aws": "ap-southeast-1", "azure": "southeastasia"},
        "悉尼": {"aws": "ap-southeast-2", "azure": "australiaeast"},
        "东京": {"aws": "ap-northeast-1", "azure": "japaneast"},
        "首尔": {"aws": "ap-northeast-2", "azure": "koreacentral"},
        "孟买": {"aws": "ap-south-1", "azure": "centralindia"},
        "香港": {"aws": "ap-east-1", "azure": "eastasia"}
    },
    "中国": {
        "北京": {"aws": "cn-north-1", "azure": "chinanorth"},
        "宁夏": {"aws": "cn-northwest-1", "azure": "chinaeast"}
    },
    "南美": {
        "圣保罗": {"aws": "sa-east-1", "azure": "brazilsouth"}
    }
}


@tool
def map_aws_to_azure_services(service_type: str) -> str:
    """
    建立AWS和Azure服务的映射关系
    
    Args:
        service_type: 服务类型关键字（ec2、ebs、s3、rds、elasticache、opensearch、elb、network）
        
    Returns:
        JSON格式的映射信息，包含aws_service、azure_service、description
    """
    try:
        service_key = service_type.lower().strip()
        
        if service_key in SERVICE_MAPPING:
            mapping = SERVICE_MAPPING[service_key]
            result = {
                "success": True,
                "service_type": service_type,
                "aws_service": mapping["aws_service"],
                "azure_service": mapping["azure_service"],
                "description": mapping["description"]
            }
            logger.info(f"成功映射服务: {service_type} -> {mapping['aws_service']} / {mapping['azure_service']}")
            return json.dumps(result, ensure_ascii=False)
        else:
            return json.dumps({
                "success": False,
                "error": f"不支持的服务类型: {service_type}",
                "supported_services": list(SERVICE_MAPPING.keys())
            }, ensure_ascii=False)
            
    except Exception as e:
        logger.error(f"映射服务失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": str(e),
            "service_type": service_type
        }, ensure_ascii=False)


@tool
def map_regions(location: str) -> str:
    """
    将地理位置映射到AWS和Azure区域代码
    
    Args:
        location: 地理位置描述（如"美国东部"、"新加坡"、"东京"）
        
    Returns:
        JSON格式的区域映射信息，包含aws_region、azure_region
    """
    try:
        location_normalized = location.strip()
        
        # 搜索所有区域
        for continent, regions in REGION_MAPPING.items():
            for region_name, region_codes in regions.items():
                if location_normalized in region_name or region_name in location_normalized:
                    result = {
                        "success": True,
                        "location": location,
                        "continent": continent,
                        "region_name": region_name,
                        "aws_region": region_codes["aws"],
                        "azure_region": region_codes["azure"]
                    }
                    logger.info(f"成功映射区域: {location} -> AWS: {region_codes['aws']}, Azure: {region_codes['azure']}")
                    return json.dumps(result, ensure_ascii=False)
        
        # 如果没有找到匹配，返回默认区域
        return json.dumps({
            "success": False,
            "error": f"未找到匹配的区域: {location}，将使用默认区域（美国东部）",
            "location": location,
            "aws_region": "us-east-1",
            "azure_region": "eastus",
            "available_regions": {
                continent: list(regions.keys())
                for continent, regions in REGION_MAPPING.items()
            }
        }, ensure_ascii=False)
        
    except Exception as e:
        logger.error(f"映射区域失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": str(e),
            "location": location,
            "aws_region": "us-east-1",
            "azure_region": "eastus"
        }, ensure_ascii=False)


@tool
def compare_pricing_across_clouds(
    aws_price_data: str,
    azure_price_data: str
) -> str:
    """
    对比AWS和Azure的价格差异
    
    Args:
        aws_price_data: AWS价格数据（JSON字符串）
        azure_price_data: Azure价格数据（JSON字符串）
        
    Returns:
        JSON格式的对比结果，包含price_difference、percentage_difference、cheaper_cloud
    """
    try:
        aws_data = json.loads(aws_price_data)
        azure_data = json.loads(azure_price_data)
        
        # 检查数据有效性
        if not aws_data.get("success") or not azure_data.get("success"):
            return json.dumps({
                "success": False,
                "error": "价格数据无效或获取失败",
                "aws_success": aws_data.get("success"),
                "azure_success": azure_data.get("success")
            }, ensure_ascii=False)
        
        # 提取年度价格进行对比
        aws_annual = aws_data.get("annual_price", 0)
        azure_annual = azure_data.get("annual_price", 0)
        
        if aws_annual == 0 or azure_annual == 0:
            return json.dumps({
                "success": False,
                "error": "价格数据为零，无法进行对比",
                "aws_annual_price": aws_annual,
                "azure_annual_price": azure_annual
            }, ensure_ascii=False)
        
        # 计算价格差异
        price_difference = abs(aws_annual - azure_annual)
        percentage_difference = (price_difference / max(aws_annual, azure_annual)) * 100
        
        # 判断哪个平台更便宜
        if aws_annual < azure_annual:
            cheaper_cloud = "AWS"
            savings = azure_annual - aws_annual
        elif azure_annual < aws_annual:
            cheaper_cloud = "Azure"
            savings = aws_annual - azure_annual
        else:
            cheaper_cloud = "相同"
            savings = 0
        
        result = {
            "success": True,
            "aws_service": aws_data.get("service"),
            "azure_service": azure_data.get("service"),
            "aws_annual_price": round(aws_annual, 2),
            "azure_annual_price": round(azure_annual, 2),
            "price_difference": round(price_difference, 2),
            "percentage_difference": round(percentage_difference, 2),
            "cheaper_cloud": cheaper_cloud,
            "annual_savings": round(savings, 2),
            "currency": "USD"
        }
        
        logger.info(f"价格对比完成: {cheaper_cloud} 更便宜，年度节省 ${savings:.2f}")
        return json.dumps(result, ensure_ascii=False)
        
    except json.JSONDecodeError as e:
        logger.error(f"解析价格数据失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": f"JSON解析错误: {str(e)}"
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"价格对比失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": str(e)
        }, ensure_ascii=False)


@tool
def calculate_annual_cost(
    items: str,
    discount_percentage: float = 0.0
) -> str:
    """
    计算年度总成本
    
    Args:
        items: 价格项目列表（JSON字符串），每个项目包含annual_price和quantity
        discount_percentage: 折扣百分比（0-100）
        
    Returns:
        JSON格式的成本计算结果，包含subtotal、discount、total
    """
    try:
        items_list = json.loads(items)
        
        if not isinstance(items_list, list):
            return json.dumps({
                "success": False,
                "error": "items参数必须是列表"
            }, ensure_ascii=False)
        
        subtotal = 0.0
        item_details = []
        
        for item in items_list:
            annual_price = item.get("annual_price", 0)
            quantity = item.get("quantity", 1)
            item_total = annual_price * quantity
            subtotal += item_total
            
            item_details.append({
                "name": item.get("name", "未知"),
                "annual_price": round(annual_price, 2),
                "quantity": quantity,
                "item_total": round(item_total, 2)
            })
        
        # 计算折扣
        discount_amount = subtotal * (discount_percentage / 100)
        total = subtotal - discount_amount
        
        result = {
            "success": True,
            "item_count": len(items_list),
            "item_details": item_details,
            "subtotal": round(subtotal, 2),
            "discount_percentage": discount_percentage,
            "discount_amount": round(discount_amount, 2),
            "total": round(total, 2),
            "currency": "USD"
        }
        
        logger.info(f"成本计算完成: 小计=${subtotal:.2f}, 折扣=${discount_amount:.2f}, 总计=${total:.2f}")
        return json.dumps(result, ensure_ascii=False)
        
    except json.JSONDecodeError as e:
        logger.error(f"解析items数据失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": f"JSON解析错误: {str(e)}"
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"计算年度成本失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": str(e)
        }, ensure_ascii=False)


@tool
def format_pricing_data(
    price: float,
    currency: str = "USD",
    decimal_places: int = 2
) -> str:
    """
    格式化价格数据为易读格式
    
    Args:
        price: 价格数值
        currency: 货币代码
        decimal_places: 小数位数
        
    Returns:
        格式化后的价格字符串（如"$1,234.56"）
    """
    try:
        # 格式化为千分位分隔符
        formatted_price = f"{price:,.{decimal_places}f}"
        
        # 添加货币符号
        currency_symbols = {
            "USD": "$",
            "EUR": "€",
            "GBP": "£",
            "CNY": "¥"
        }
        
        symbol = currency_symbols.get(currency, currency)
        result = f"{symbol}{formatted_price}"
        
        return result
        
    except Exception as e:
        logger.error(f"格式化价格失败: {str(e)}")
        return str(price)


@tool
def generate_excel_report(
    report_data: str,
    output_path: str = None
) -> str:
    """
    生成包含三个Sheet的Excel报告
    
    Args:
        report_data: 报告数据（JSON字符串），包含aws_items、azure_items、comparison_items
        output_path: 输出文件路径，如果为None则自动生成
        
    Returns:
        JSON格式的生成结果，包含file_path、success
    """
    try:
        data = json.loads(report_data)
        
        # 如果未指定输出路径，自动生成
        if output_path is None:
            cache_dir = Path(".cache/multi_cloud_pricing_comparison_agent/reports")
            cache_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(cache_dir / f"多云报价对比_{timestamp}.xlsx")
        
        # 创建工作簿
        wb = Workbook()
        
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='微软雅黑', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Sheet 1: AWS报价表
        ws_aws = wb.active
        ws_aws.title = "AWS报价"
        
        aws_headers = ["序号", "项目", "服务名称", "配置", "数量", "单价(USD/年)", "1年总价(USD)", "备注"]
        ws_aws.append(aws_headers)
        
        # 设置表头样式
        for col_num, header in enumerate(aws_headers, 1):
            cell = ws_aws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 填充AWS数据（支持中英文字段名）
        aws_items = data.get("aws_items", [])
        for idx, item in enumerate(aws_items, 1):
            # 兼容中英文字段名
            project_name = item.get("project_name") or item.get("项目", "")
            service_name = item.get("service_name") or item.get("服务名称", "")
            configuration = item.get("configuration") or item.get("配置", "")
            quantity = item.get("quantity") or item.get("数量", 1)
            # 优先使用单价，如果没有则用年总价除以数量
            annual_price = item.get("annual_price") or item.get("单价") or (item.get("年总价", 0) / max(quantity, 1))
            notes = item.get("notes") or item.get("备注", "")
            
            row_data = [
                idx,
                project_name,
                service_name,
                configuration,
                quantity,
                annual_price,
                annual_price * quantity,
                notes
            ]
            ws_aws.append(row_data)
            
            # 设置单元格样式
            for col_num in range(1, len(row_data) + 1):
                cell = ws_aws.cell(row=idx + 1, column=col_num)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border
        
        # 添加总计行（支持中英文字段名）
        total_row = len(aws_items) + 2
        ws_aws.cell(row=total_row, column=6, value="总计:")
        
        def calc_item_total(item):
            quantity = item.get("quantity") or item.get("数量", 1)
            annual_price = item.get("annual_price") or item.get("单价") or (item.get("年总价", 0) / max(quantity, 1))
            return annual_price * quantity
        
        ws_aws.cell(row=total_row, column=7, value=sum(calc_item_total(item) for item in aws_items))
        ws_aws.cell(row=total_row, column=6).font = Font(bold=True)
        ws_aws.cell(row=total_row, column=7).font = Font(bold=True)
        
        # 调整列宽
        column_widths = [8, 20, 25, 30, 10, 18, 18, 40]
        for col_num, width in enumerate(column_widths, 1):
            ws_aws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Sheet 2: Azure报价表
        ws_azure = wb.create_sheet("Azure报价")
        
        azure_headers = ["序号", "项目", "服务名称", "配置", "数量", "单价(USD/年)", "1年总价(USD)", "备注"]
        ws_azure.append(azure_headers)
        
        # 设置表头样式
        for col_num, header in enumerate(azure_headers, 1):
            cell = ws_azure.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 填充Azure数据（支持中英文字段名）
        azure_items = data.get("azure_items", [])
        for idx, item in enumerate(azure_items, 1):
            # 兼容中英文字段名
            project_name = item.get("project_name") or item.get("项目", "")
            service_name = item.get("service_name") or item.get("服务名称", "")
            configuration = item.get("configuration") or item.get("配置", "")
            quantity = item.get("quantity") or item.get("数量", 1)
            # 优先使用单价，如果没有则用年总价除以数量
            annual_price = item.get("annual_price") or item.get("单价") or (item.get("年总价", 0) / max(quantity, 1))
            notes = item.get("notes") or item.get("备注", "")
            
            row_data = [
                idx,
                project_name,
                service_name,
                configuration,
                quantity,
                annual_price,
                annual_price * quantity,
                notes
            ]
            ws_azure.append(row_data)
            
            # 设置单元格样式
            for col_num in range(1, len(row_data) + 1):
                cell = ws_azure.cell(row=idx + 1, column=col_num)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border
        
        # 添加总计行（支持中英文字段名）
        total_row = len(azure_items) + 2
        ws_azure.cell(row=total_row, column=6, value="总计:")
        ws_azure.cell(row=total_row, column=7, value=sum(calc_item_total(item) for item in azure_items))
        ws_azure.cell(row=total_row, column=6).font = Font(bold=True)
        ws_azure.cell(row=total_row, column=7).font = Font(bold=True)
        
        # 调整列宽
        for col_num, width in enumerate(column_widths, 1):
            ws_azure.column_dimensions[get_column_letter(col_num)].width = width
        
        # Sheet 3: 对比总结表
        ws_comparison = wb.create_sheet("对比总结")
        
        comparison_headers = ["序号", "服务类型", "AWS总价(USD/年)", "Azure总价(USD/年)", "价格差异(USD)", "差异百分比(%)", "推荐平台", "备注"]
        ws_comparison.append(comparison_headers)
        
        # 设置表头样式
        for col_num, header in enumerate(comparison_headers, 1):
            cell = ws_comparison.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 填充对比数据（支持中英文字段名和多种价格格式）
        comparison_items = data.get("comparison_items", [])
        
        def parse_price(price_str):
            """解析价格字符串，支持 '$1,234.56' 或数字格式"""
            if isinstance(price_str, (int, float)):
                return float(price_str)
            if isinstance(price_str, str):
                # 移除货币符号和逗号
                clean_str = price_str.replace("$", "").replace(",", "").strip()
                try:
                    return float(clean_str)
                except ValueError:
                    return 0
            return 0
        
        for idx, item in enumerate(comparison_items, 1):
            # 兼容中英文字段名
            service_type = item.get("service_type") or item.get("项目", "")
            
            # 解析价格（支持字符串格式如 "$1,234.56"）
            aws_price_raw = item.get("aws_annual_price") or item.get("AWS价格", 0)
            azure_price_raw = item.get("azure_annual_price") or item.get("Azure价格", 0)
            aws_price = parse_price(aws_price_raw)
            azure_price = parse_price(azure_price_raw)
            
            # 解析差异和百分比
            diff_raw = item.get("price_difference") or item.get("差异", 0)
            price_diff = parse_price(diff_raw) if diff_raw else abs(aws_price - azure_price)
            
            # 计算百分比
            percentage_str = item.get("percentage_difference") or item.get("节省百分比", "")
            if isinstance(percentage_str, str) and "%" in percentage_str:
                # 提取百分比数字
                import re
                match = re.search(r'(\d+\.?\d*)', percentage_str)
                percentage_diff = float(match.group(1)) if match else 0
            elif isinstance(percentage_str, (int, float)):
                percentage_diff = float(percentage_str)
            else:
                percentage_diff = (price_diff / max(aws_price, azure_price) * 100) if max(aws_price, azure_price) > 0 else 0
            
            # 判断推荐平台
            notes_str = item.get("notes") or item.get("节省百分比", "")
            if "AWS" in str(notes_str) and "便宜" in str(notes_str):
                recommended = "AWS"
            elif "Azure" in str(notes_str) and "便宜" in str(notes_str):
                recommended = "Azure"
            elif aws_price < azure_price:
                recommended = "AWS"
            elif azure_price < aws_price:
                recommended = "Azure"
            else:
                recommended = "相同"
            
            row_data = [
                idx,
                service_type,
                aws_price,
                azure_price,
                price_diff,
                round(percentage_diff, 2),
                recommended,
                item.get("notes") or item.get("节省百分比", "")
            ]
            ws_comparison.append(row_data)
            
            # 设置单元格样式
            for col_num in range(1, len(row_data) + 1):
                cell = ws_comparison.cell(row=idx + 1, column=col_num)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border
        
        # 添加总计行（支持中英文字段名）
        # 注意：如果comparison_items最后一行已经是总计，则不重复添加
        total_row = len(comparison_items) + 2
        
        # 检查最后一项是否是总计行
        last_item = comparison_items[-1] if comparison_items else {}
        last_service_type = last_item.get("service_type") or last_item.get("项目", "")
        
        if "总计" not in str(last_service_type):
            # 计算总计
            aws_total = sum(
                parse_price(item.get("aws_annual_price") or item.get("AWS价格", 0))
                for item in comparison_items
            )
            azure_total = sum(
                parse_price(item.get("azure_annual_price") or item.get("Azure价格", 0))
                for item in comparison_items
            )
            
            ws_comparison.cell(row=total_row, column=2, value="总计:")
            ws_comparison.cell(row=total_row, column=3, value=aws_total)
            ws_comparison.cell(row=total_row, column=4, value=azure_total)
            ws_comparison.cell(row=total_row, column=2).font = Font(bold=True)
            ws_comparison.cell(row=total_row, column=3).font = Font(bold=True)
            ws_comparison.cell(row=total_row, column=4).font = Font(bold=True)
        
        # 调整列宽
        comparison_widths = [8, 20, 20, 20, 18, 18, 15, 30]
        for col_num, width in enumerate(comparison_widths, 1):
            ws_comparison.column_dimensions[get_column_letter(col_num)].width = width
        
        # 保存工作簿
        wb.save(output_path)
        
        result = {
            "success": True,
            "file_path": output_path,
            "sheets": ["AWS报价", "Azure报价", "对比总结"],
            "aws_item_count": len(aws_items),
            "azure_item_count": len(azure_items),
            "comparison_item_count": len(comparison_items),
            "generated_at": datetime.now().isoformat()
        }
        
        logger.info(f"Excel报告生成成功: {output_path}")
        return json.dumps(result, ensure_ascii=False)
        
    except json.JSONDecodeError as e:
        logger.error(f"解析报告数据失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": f"JSON解析错误: {str(e)}"
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"生成Excel报告失败: {str(e)}")
        return json.dumps({
            "success": False,
            "error": str(e)
        }, ensure_ascii=False)
