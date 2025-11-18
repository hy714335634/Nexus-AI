#!/usr/bin/env python3
"""
Excel数据处理工具模块

提供Excel文件读取、数据分析和缓存管理功能。
所有工具函数遵循Strands框架规范，使用@tool装饰器。
"""

import json
import os
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from strands import tool


@tool
def excel_data_reader(
    file_path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 0,
    skip_rows: Optional[List[int]] = None
) -> str:
    """
    Excel文件读取工具
    
    功能：读取.xlsx和.xls格式的Excel文件，支持多工作表识别、空值处理、异常值记录
    
    Args:
        file_path (str): Excel文件路径
        sheet_name (Optional[str]): 工作表名称，如果不提供则读取第一个工作表
        header_row (int): 表头行号，默认为0（第一行）
        skip_rows (Optional[List[int]]): 要跳过的行号列表
        
    Returns:
        str: JSON格式的读取结果，包含DataFrame数据、元数据和异常值信息
    """
    try:
        # 验证文件路径
        if not os.path.exists(file_path):
            return json.dumps({
                "status": "error",
                "message": f"文件不存在: {file_path}",
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        # 验证文件格式
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in ['.xlsx', '.xls']:
            return json.dumps({
                "status": "error",
                "message": f"不支持的文件格式: {file_ext}，仅支持.xlsx和.xls",
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        # 获取所有工作表名称
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        all_sheets = workbook.sheetnames
        workbook.close()
        
        # 确定要读取的工作表
        target_sheet = sheet_name if sheet_name else all_sheets[0]
        
        if sheet_name and sheet_name not in all_sheets:
            return json.dumps({
                "status": "error",
                "message": f"工作表'{sheet_name}'不存在",
                "available_sheets": all_sheets,
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        # 读取Excel数据
        df = pd.read_excel(
            file_path,
            sheet_name=target_sheet,
            header=header_row,
            skiprows=skip_rows
        )
        
        # 处理空值
        null_counts = df.isnull().sum().to_dict()
        null_percentages = (df.isnull().sum() / len(df) * 100).to_dict()
        
        # 识别异常值（使用IQR方法）
        outliers = {}
        for col in df.select_dtypes(include=[np.number]).columns:
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            
            outlier_indices = df[(df[col] < lower_bound) | (df[col] > upper_bound)].index.tolist()
            if outlier_indices:
                outliers[col] = {
                    "count": len(outlier_indices),
                    "indices": outlier_indices[:10],  # 最多返回前10个
                    "lower_bound": float(lower_bound),
                    "upper_bound": float(upper_bound)
                }
        
        # 数据类型信息
        dtypes_info = {col: str(dtype) for col, dtype in df.dtypes.items()}
        
        # 基础统计信息
        numeric_stats = {}
        for col in df.select_dtypes(include=[np.number]).columns:
            numeric_stats[col] = {
                "count": int(df[col].count()),
                "mean": float(df[col].mean()) if not pd.isna(df[col].mean()) else None,
                "std": float(df[col].std()) if not pd.isna(df[col].std()) else None,
                "min": float(df[col].min()) if not pd.isna(df[col].min()) else None,
                "25%": float(df[col].quantile(0.25)) if not pd.isna(df[col].quantile(0.25)) else None,
                "50%": float(df[col].quantile(0.50)) if not pd.isna(df[col].quantile(0.50)) else None,
                "75%": float(df[col].quantile(0.75)) if not pd.isna(df[col].quantile(0.75)) else None,
                "max": float(df[col].max()) if not pd.isna(df[col].max()) else None
            }
        
        # 分类列信息
        categorical_info = {}
        for col in df.select_dtypes(include=['object']).columns:
            value_counts = df[col].value_counts().head(10).to_dict()
            categorical_info[col] = {
                "unique_count": int(df[col].nunique()),
                "top_values": {str(k): int(v) for k, v in value_counts.items()},
                "null_count": int(df[col].isnull().sum())
            }
        
        # 转换DataFrame为JSON格式
        data_json = df.to_dict(orient='records')
        
        # 构建返回结果
        result = {
            "status": "success",
            "file_info": {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "file_size_bytes": os.path.getsize(file_path),
                "sheet_name": target_sheet,
                "all_sheets": all_sheets,
                "read_time": datetime.now().isoformat()
            },
            "data_shape": {
                "rows": len(df),
                "columns": len(df.columns),
                "total_cells": len(df) * len(df.columns)
            },
            "columns": {
                "names": df.columns.tolist(),
                "data_types": dtypes_info
            },
            "data": data_json,
            "metadata": {
                "null_counts": {k: int(v) for k, v in null_counts.items()},
                "null_percentages": {k: float(v) for k, v in null_percentages.items()},
                "numeric_statistics": numeric_stats,
                "categorical_info": categorical_info,
                "outliers": outliers
            },
            "data_quality": {
                "completeness_score": float(100 - (df.isnull().sum().sum() / (len(df) * len(df.columns)) * 100)),
                "total_nulls": int(df.isnull().sum().sum()),
                "total_outliers": sum(info["count"] for info in outliers.values())
            }
        }
        
        return json.dumps(result, ensure_ascii=False, indent=2)
        
    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": f"读取Excel文件失败: {str(e)}",
            "error_type": type(e).__name__,
            "timestamp": datetime.now().isoformat()
        }, ensure_ascii=False, indent=2)


@tool
def data_analyzer(
    data_json: str,
    analysis_dimensions: Optional[List[str]] = None,
    include_correlation: bool = True,
    include_trends: bool = True
) -> str:
    """
    数据分析工具
    
    功能：进行深度数据分析，包括统计分析、趋势分析、相关性分析和异常检测
    
    Args:
        data_json (str): JSON格式的DataFrame数据（通常来自excel_data_reader）
        analysis_dimensions (Optional[List[str]]): 要分析的列名列表，如果不提供则分析所有数值列
        include_correlation (bool): 是否包含相关性分析，默认True
        include_trends (bool): 是否包含趋势分析，默认True
        
    Returns:
        str: JSON格式的分析结果
    """
    try:
        # 解析输入数据
        input_data = json.loads(data_json)
        
        # 提取DataFrame数据
        if "data" in input_data:
            df_data = input_data["data"]
        else:
            df_data = input_data
        
        df = pd.DataFrame(df_data)
        
        if df.empty:
            return json.dumps({
                "status": "error",
                "message": "数据为空，无法进行分析",
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        # 确定分析维度
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        if analysis_dimensions:
            numeric_columns = [col for col in analysis_dimensions if col in numeric_columns]
        
        if not numeric_columns:
            return json.dumps({
                "status": "error",
                "message": "没有可分析的数值列",
                "available_columns": df.columns.tolist(),
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        # 基础统计分析
        statistical_analysis = {}
        for col in numeric_columns:
            col_data = df[col].dropna()
            if len(col_data) == 0:
                continue
                
            statistical_analysis[col] = {
                "count": int(len(col_data)),
                "mean": float(col_data.mean()),
                "median": float(col_data.median()),
                "std": float(col_data.std()),
                "variance": float(col_data.var()),
                "min": float(col_data.min()),
                "max": float(col_data.max()),
                "range": float(col_data.max() - col_data.min()),
                "q1": float(col_data.quantile(0.25)),
                "q3": float(col_data.quantile(0.75)),
                "iqr": float(col_data.quantile(0.75) - col_data.quantile(0.25)),
                "skewness": float(col_data.skew()),
                "kurtosis": float(col_data.kurtosis()),
                "coefficient_of_variation": float(col_data.std() / col_data.mean() * 100) if col_data.mean() != 0 else None
            }
        
        # 相关性分析
        correlation_matrix = {}
        if include_correlation and len(numeric_columns) > 1:
            corr = df[numeric_columns].corr()
            correlation_matrix = {
                "matrix": corr.to_dict(),
                "high_correlations": []
            }
            
            # 识别高相关性对
            for i in range(len(corr.columns)):
                for j in range(i+1, len(corr.columns)):
                    corr_value = corr.iloc[i, j]
                    if abs(corr_value) > 0.7:  # 高相关性阈值
                        correlation_matrix["high_correlations"].append({
                            "variable1": corr.columns[i],
                            "variable2": corr.columns[j],
                            "correlation": float(corr_value),
                            "strength": "strong" if abs(corr_value) > 0.9 else "moderate"
                        })
        
        # 趋势分析（时间序列识别）
        trend_analysis = {}
        if include_trends:
            # 尝试识别时间列
            time_columns = []
            for col in df.columns:
                try:
                    # 使用 errors='coerce' 避免警告，如果无法解析则返回 NaT
                    parsed_dates = pd.to_datetime(df[col], errors='coerce')
                    # 检查是否有足够多的有效日期（至少50%是有效日期）
                    if parsed_dates.notna().sum() / len(parsed_dates) > 0.5:
                        time_columns.append(col)
                except:
                    continue
            
            if time_columns:
                time_col = time_columns[0]
                df_sorted = df.sort_values(by=time_col)
                
                for col in numeric_columns:
                    if col != time_col:
                        values = df_sorted[col].dropna()
                        if len(values) > 3:
                            # 计算趋势（线性回归斜率）
                            x = np.arange(len(values))
                            y = values.values
                            slope = np.polyfit(x, y, 1)[0]
                            
                            trend_analysis[col] = {
                                "trend_direction": "increasing" if slope > 0 else "decreasing",
                                "slope": float(slope),
                                "start_value": float(values.iloc[0]),
                                "end_value": float(values.iloc[-1]),
                                "change_percentage": float((values.iloc[-1] - values.iloc[0]) / values.iloc[0] * 100) if values.iloc[0] != 0 else None,
                                "time_column": time_col
                            }
        
        # 异常值检测
        anomaly_detection = {}
        for col in numeric_columns:
            col_data = df[col].dropna()
            if len(col_data) < 4:
                continue
            
            Q1 = col_data.quantile(0.25)
            Q3 = col_data.quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            
            outliers = df[(df[col] < lower_bound) | (df[col] > upper_bound)]
            if len(outliers) > 0:
                anomaly_detection[col] = {
                    "count": len(outliers),
                    "percentage": float(len(outliers) / len(col_data) * 100),
                    "outlier_indices": outliers.index.tolist()[:20],  # 最多返回前20个
                    "outlier_values": [float(v) for v in outliers[col].tolist()[:20]],
                    "bounds": {
                        "lower": float(lower_bound),
                        "upper": float(upper_bound)
                    }
                }
        
        # 分组统计（如果存在分类列）
        groupby_analysis = {}
        categorical_columns = df.select_dtypes(include=['object']).columns.tolist()
        if categorical_columns and numeric_columns:
            # 选择第一个分类列和第一个数值列进行分组统计
            cat_col = categorical_columns[0]
            num_col = numeric_columns[0]
            
            grouped = df.groupby(cat_col)[num_col].agg(['count', 'mean', 'std', 'min', 'max'])
            groupby_analysis = {
                "groupby_column": cat_col,
                "value_column": num_col,
                "groups": grouped.to_dict(orient='index')
            }
        
        # 构建分析结果
        result = {
            "status": "success",
            "analysis_time": datetime.now().isoformat(),
            "data_shape": {
                "rows": len(df),
                "columns": len(df.columns)
            },
            "analyzed_columns": numeric_columns,
            "statistical_analysis": statistical_analysis,
            "correlation_analysis": correlation_matrix if include_correlation else {},
            "trend_analysis": trend_analysis if include_trends else {},
            "anomaly_detection": anomaly_detection,
            "groupby_analysis": groupby_analysis,
            "data_insights": {
                "total_numeric_columns": len(numeric_columns),
                "total_categorical_columns": len(categorical_columns),
                "columns_with_outliers": len(anomaly_detection),
                "high_correlation_pairs": len(correlation_matrix.get("high_correlations", [])) if include_correlation else 0
            }
        }
        
        return json.dumps(result, ensure_ascii=False, indent=2)
        
    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": f"数据分析失败: {str(e)}",
            "error_type": type(e).__name__,
            "timestamp": datetime.now().isoformat()
        }, ensure_ascii=False, indent=2)


@tool
def cache_manager(
    operation: str,
    session_id: str,
    file_info: Optional[Dict[str, Any]] = None,
    cache_base_dir: str = ".cache/excel_report_generator"
) -> str:
    """
    缓存管理工具
    
    功能：管理生成的图表和报告缓存，支持创建会话目录、文件复用检测、过期清理
    
    Args:
        operation (str): 操作类型 (create_session/check_cache/save_cache/list_cache/clean_cache)
        session_id (str): 会话ID
        file_info (Optional[Dict[str, Any]]): 文件信息字典（用于save_cache和check_cache）
        cache_base_dir (str): 缓存基础目录，默认为.cache/excel_report_generator
        
    Returns:
        str: JSON格式的操作结果
    """
    try:
        cache_dir = Path(cache_base_dir)
        session_dir = cache_dir / session_id
        
        if operation == "create_session":
            # 创建会话目录
            session_dir.mkdir(parents=True, exist_ok=True)
            
            # 创建会话元数据
            metadata = {
                "session_id": session_id,
                "created_at": datetime.now().isoformat(),
                "status": "active"
            }
            
            metadata_file = session_dir / "session_metadata.json"
            with open(metadata_file, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)
            
            return json.dumps({
                "status": "success",
                "operation": "create_session",
                "session_id": session_id,
                "session_path": str(session_dir),
                "metadata_file": str(metadata_file),
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        elif operation == "check_cache":
            # 检查缓存是否存在
            if not file_info or "file_hash" not in file_info:
                return json.dumps({
                    "status": "error",
                    "message": "file_info必须包含file_hash字段",
                    "timestamp": datetime.now().isoformat()
                }, ensure_ascii=False, indent=2)
            
            file_hash = file_info["file_hash"]
            cached_files = list(session_dir.glob(f"*{file_hash}*")) if session_dir.exists() else []
            
            return json.dumps({
                "status": "success",
                "operation": "check_cache",
                "cache_exists": len(cached_files) > 0,
                "cached_files": [str(f) for f in cached_files],
                "file_count": len(cached_files),
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        elif operation == "save_cache":
            # 保存缓存信息
            if not file_info:
                return json.dumps({
                    "status": "error",
                    "message": "file_info不能为空",
                    "timestamp": datetime.now().isoformat()
                }, ensure_ascii=False, indent=2)
            
            session_dir.mkdir(parents=True, exist_ok=True)
            
            # 保存缓存记录
            cache_record_file = session_dir / "cache_records.json"
            cache_records = []
            
            if cache_record_file.exists():
                with open(cache_record_file, 'r', encoding='utf-8') as f:
                    cache_records = json.load(f)
            
            cache_records.append({
                **file_info,
                "cached_at": datetime.now().isoformat()
            })
            
            with open(cache_record_file, 'w', encoding='utf-8') as f:
                json.dump(cache_records, f, ensure_ascii=False, indent=2)
            
            return json.dumps({
                "status": "success",
                "operation": "save_cache",
                "cache_record_file": str(cache_record_file),
                "total_cached_files": len(cache_records),
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        elif operation == "list_cache":
            # 列出缓存文件
            if not session_dir.exists():
                return json.dumps({
                    "status": "success",
                    "operation": "list_cache",
                    "cached_files": [],
                    "total_files": 0,
                    "timestamp": datetime.now().isoformat()
                }, ensure_ascii=False, indent=2)
            
            cache_record_file = session_dir / "cache_records.json"
            cache_records = []
            
            if cache_record_file.exists():
                with open(cache_record_file, 'r', encoding='utf-8') as f:
                    cache_records = json.load(f)
            
            return json.dumps({
                "status": "success",
                "operation": "list_cache",
                "session_id": session_id,
                "cached_files": cache_records,
                "total_files": len(cache_records),
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        elif operation == "clean_cache":
            # 清理过期缓存
            if not session_dir.exists():
                return json.dumps({
                    "status": "success",
                    "operation": "clean_cache",
                    "message": "会话目录不存在，无需清理",
                    "timestamp": datetime.now().isoformat()
                }, ensure_ascii=False, indent=2)
            
            # 删除会话目录及其所有内容
            import shutil
            shutil.rmtree(session_dir)
            
            return json.dumps({
                "status": "success",
                "operation": "clean_cache",
                "session_id": session_id,
                "cleaned_path": str(session_dir),
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
        else:
            return json.dumps({
                "status": "error",
                "message": f"不支持的操作: {operation}",
                "supported_operations": ["create_session", "check_cache", "save_cache", "list_cache", "clean_cache"],
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False, indent=2)
        
    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": f"缓存管理失败: {str(e)}",
            "error_type": type(e).__name__,
            "timestamp": datetime.now().isoformat()
        }, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    print("🧪 测试Excel数据处理工具...")
    
    # 这里可以添加测试代码
    print("✅ Excel数据处理工具模块加载成功！")
