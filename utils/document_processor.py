"""
Document processor for handling Excel and Word files.

This module provides the DocumentProcessor class that handles:
- Excel file processing using pandas and openpyxl
- Word document processing using python-docx
- Structured document to text conversion
- Integration with multimodal model service for content processing
"""

import os
import time
import logging
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import load_workbook
# from docx import Document
from io import StringIO

from models.interfaces import FileProcessor
from models.data_models import FileMetadata, ProcessedContent
from models.exceptions import FileProcessingError
from utils.multimodal_model_service import MultimodalModelService


class DocumentProcessor(FileProcessor):
    """
    Processor for handling structured documents (Excel, Word, CSV).
    
    This processor can handle:
    - Excel files (.xlsx, .xls) using pandas and openpyxl
    - Word documents (.docx) using python-docx
    - CSV files using pandas
    
    The processor extracts structured content and converts it to text format,
    then uses the multimodal model service for final formatting and analysis.
    """
    
    def __init__(self, multimodal_service: Optional[MultimodalModelService] = None):
        """
        Initialize the document processor.
        
        Args:
            multimodal_service: Optional multimodal model service instance.
                               If None, a new instance will be created.
        """
        self.logger = logging.getLogger(__name__)
        self.multimodal_service = multimodal_service or MultimodalModelService()
        
        # Supported file types
        self._supported_types = ['xlsx', 'xls', 'docx', 'csv']
        
        # Excel processing settings
        self.max_rows_per_sheet = 10000  # Limit to prevent memory issues
        self.max_sheets = 20  # Limit number of sheets to process
        
    def can_process(self, file_type: str) -> bool:
        """
        Check if this processor can handle the given file type.
        
        Args:
            file_type: File extension (e.g., 'xlsx', 'docx')
            
        Returns:
            True if this processor can handle the file type
        """
        return file_type.lower() in self._supported_types
    
    def get_supported_types(self) -> List[str]:
        """
        Get list of file types supported by this processor.
        
        Returns:
            List of supported file extensions
        """
        return self._supported_types.copy()
    
    def process(self, file_metadata: FileMetadata) -> ProcessedContent:
        """
        Process a document file and return the processed content.
        
        Args:
            file_metadata: Metadata about the file to process
            
        Returns:
            ProcessedContent object containing the processing results
            
        Raises:
            FileProcessingError: If processing fails
        """
        start_time = time.time()
        
        try:
            self.logger.info(f"Processing document: {file_metadata.original_name}")
            
            # Download file from S3 if needed (assuming file is accessible via s3_url)
            # For now, we'll assume the file is locally accessible
            # In a real implementation, you'd download from S3 first
            file_path = self._get_local_file_path(file_metadata)
            
            # Extract content based on file type
            extracted_content = self._extract_document_content(file_path, file_metadata.file_type)
            
            # Process with multimodal model
            file_info = {
                'original_name': file_metadata.original_name,
                'file_type': file_metadata.file_type,
                'file_size': file_metadata.file_size,
                'upload_time': file_metadata.upload_time.isoformat()
            }
            
            processed_text = self.multimodal_service.process_text_content(
                extracted_content, 
                file_info
            )
            
            processing_time = time.time() - start_time
            
            # Extract metadata from the processing
            metadata = {
                'extraction_method': self._get_extraction_method(file_metadata.file_type),
                'content_length': len(extracted_content),
                'processed_length': len(processed_text),
                'file_type': file_metadata.file_type
            }
            
            # Add specific metadata based on file type
            if file_metadata.file_type.lower() in ['xlsx', 'xls']:
                metadata.update(self._get_excel_metadata(file_path))
            elif file_metadata.file_type.lower() == 'docx':
                metadata.update(self._get_word_metadata(file_path))
            elif file_metadata.file_type.lower() == 'csv':
                metadata.update(self._get_csv_metadata(file_path))
            
            return ProcessedContent(
                file_id=file_metadata.file_id,
                file_name=file_metadata.original_name,
                content_type='document',
                processed_text=processed_text,
                metadata=metadata,
                processing_time=processing_time,
                success=True
            )
            
        except Exception as e:
            processing_time = time.time() - start_time
            error_msg = f"Failed to process document {file_metadata.original_name}: {str(e)}"
            self.logger.error(error_msg)
            
            return ProcessedContent(
                file_id=file_metadata.file_id,
                file_name=file_metadata.original_name,
                content_type='document',
                processed_text='',
                metadata={'error_type': type(e).__name__},
                processing_time=processing_time,
                success=False,
                error_message=error_msg
            )
    
    def _get_local_file_path(self, file_metadata: FileMetadata) -> str:
        """
        Get local file path for processing.
        
        In a real implementation, this would download the file from S3.
        For now, we'll assume files are in a local temp directory.
        
        Args:
            file_metadata: File metadata containing S3 URL
            
        Returns:
            Local file path
        """
        # This is a placeholder implementation
        # In reality, you'd download from S3 using the s3_url
        temp_dir = "/tmp/multimodal_parser"
        os.makedirs(temp_dir, exist_ok=True)
        return os.path.join(temp_dir, f"{file_metadata.file_id}_{file_metadata.original_name}")
    
    def _extract_document_content(self, file_path: str, file_type: str) -> str:
        """
        Extract content from document based on file type.
        
        Args:
            file_path: Path to the document file
            file_type: Type of the file (xlsx, docx, csv)
            
        Returns:
            Extracted text content
            
        Raises:
            FileProcessingError: If extraction fails
        """
        file_type = file_type.lower()
        
        try:
            if file_type in ['xlsx', 'xls']:
                return self._extract_excel_content(file_path)
            elif file_type == 'docx':
                return self._extract_word_content(file_path)
            elif file_type == 'csv':
                return self._extract_csv_content(file_path)
            else:
                raise FileProcessingError(
                    f"Unsupported file type: {file_type}",
                    error_code="UNSUPPORTED_FILE_TYPE",
                    context={"file_type": file_type, "file_path": file_path}
                )
                
        except FileProcessingError:
            raise
        except Exception as e:
            raise FileProcessingError(
                f"Failed to extract content from {file_type} file: {str(e)}",
                error_code="CONTENT_EXTRACTION_ERROR",
                context={"file_type": file_type, "file_path": file_path, "error": str(e)}
            )
    
    def _extract_excel_content(self, file_path: str) -> str:
        """
        Extract content from Excel file using pandas and openpyxl.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Extracted content as formatted text
        """
        content_parts = []
        
        try:
            # First, get sheet names using openpyxl for better metadata
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames[:self.max_sheets]  # Limit sheets
            workbook.close()
            
            content_parts.append(f"Excel文件包含 {len(sheet_names)} 个工作表")
            content_parts.append("=" * 50)
            
            # Process each sheet with pandas
            for i, sheet_name in enumerate(sheet_names, 1):
                try:
                    # Read sheet with pandas
                    df = pd.read_excel(
                        file_path, 
                        sheet_name=sheet_name,
                        nrows=self.max_rows_per_sheet
                    )
                    
                    content_parts.append(f"\n工作表 {i}: {sheet_name}")
                    content_parts.append("-" * 30)
                    
                    # Add basic info
                    content_parts.append(f"行数: {len(df)}")
                    content_parts.append(f"列数: {len(df.columns)}")
                    content_parts.append(f"列名: {', '.join(df.columns.astype(str))}")
                    
                    # Convert DataFrame to string representation
                    if not df.empty:
                        # Show first few rows and summary
                        content_parts.append("\n前几行数据:")
                        content_parts.append(df.head(10).to_string(index=False))
                        
                        # Add data types info
                        content_parts.append(f"\n数据类型:")
                        for col, dtype in df.dtypes.items():
                            content_parts.append(f"  {col}: {dtype}")
                        
                        # Add basic statistics for numeric columns
                        numeric_cols = df.select_dtypes(include=['number']).columns
                        if len(numeric_cols) > 0:
                            content_parts.append(f"\n数值列统计:")
                            stats = df[numeric_cols].describe()
                            content_parts.append(stats.to_string())
                    else:
                        content_parts.append("工作表为空")
                        
                except Exception as e:
                    content_parts.append(f"\n处理工作表 '{sheet_name}' 时出错: {str(e)}")
                    
        except Exception as e:
            raise FileProcessingError(
                f"Failed to process Excel file: {str(e)}",
                error_code="EXCEL_PROCESSING_ERROR",
                context={"file_path": file_path, "error": str(e)}
            )
        
        return "\n".join(content_parts)
    
    def _extract_word_content(self, file_path: str) -> str:
        """
        Extract content from Word document using python-docx.
        
        Args:
            file_path: Path to the Word document
            
        Returns:
            Extracted content as formatted text
        """
        content_parts = []
        
        try:
            doc = Document(file_path)
            
            content_parts.append("Word文档内容")
            content_parts.append("=" * 50)
            
            # Extract paragraphs
            paragraph_count = 0
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():  # Skip empty paragraphs
                    paragraph_count += 1
                    # Preserve some formatting information
                    text = paragraph.text.strip()
                    
                    # Check if it looks like a heading (simple heuristic)
                    if len(text) < 100 and not text.endswith('.'):
                        content_parts.append(f"\n## {text}")
                    else:
                        content_parts.append(f"\n{text}")
            
            # Extract tables
            table_count = 0
            for table in doc.tables:
                table_count += 1
                content_parts.append(f"\n\n表格 {table_count}:")
                content_parts.append("-" * 20)
                
                # Convert table to text format
                table_data = []
                for row in table.rows:
                    row_data = []
                    for cell in row.cells:
                        row_data.append(cell.text.strip())
                    table_data.append(row_data)
                
                # Format as simple table
                if table_data:
                    # Use first row as headers if it looks like headers
                    headers = table_data[0]
                    content_parts.append(" | ".join(headers))
                    content_parts.append(" | ".join(["-" * len(h) for h in headers]))
                    
                    for row in table_data[1:]:
                        content_parts.append(" | ".join(row))
            
            # Add document statistics
            content_parts.append(f"\n\n文档统计:")
            content_parts.append(f"段落数: {paragraph_count}")
            content_parts.append(f"表格数: {table_count}")
            
        except Exception as e:
            raise FileProcessingError(
                f"Failed to process Word document: {str(e)}",
                error_code="WORD_PROCESSING_ERROR",
                context={"file_path": file_path, "error": str(e)}
            )
        
        return "\n".join(content_parts)
    
    def _extract_csv_content(self, file_path: str) -> str:
        """
        Extract content from CSV file using pandas.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Extracted content as formatted text
        """
        content_parts = []
        
        try:
            # Try different encodings
            encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(
                        file_path, 
                        encoding=encoding,
                        nrows=self.max_rows_per_sheet
                    )
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise FileProcessingError(
                    "Failed to read CSV file with any supported encoding",
                    error_code="CSV_ENCODING_ERROR",
                    context={"file_path": file_path, "tried_encodings": encodings}
                )
            
            content_parts.append("CSV文件内容")
            content_parts.append("=" * 50)
            
            # Add basic info
            content_parts.append(f"行数: {len(df)}")
            content_parts.append(f"列数: {len(df.columns)}")
            content_parts.append(f"列名: {', '.join(df.columns.astype(str))}")
            
            # Show data preview
            if not df.empty:
                content_parts.append("\n数据预览:")
                content_parts.append(df.head(10).to_string(index=False))
                
                # Add data types
                content_parts.append(f"\n数据类型:")
                for col, dtype in df.dtypes.items():
                    content_parts.append(f"  {col}: {dtype}")
                
                # Add basic statistics for numeric columns
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    content_parts.append(f"\n数值列统计:")
                    stats = df[numeric_cols].describe()
                    content_parts.append(stats.to_string())
            else:
                content_parts.append("CSV文件为空")
                
        except FileProcessingError:
            raise
        except Exception as e:
            raise FileProcessingError(
                f"Failed to process CSV file: {str(e)}",
                error_code="CSV_PROCESSING_ERROR",
                context={"file_path": file_path, "error": str(e)}
            )
        
        return "\n".join(content_parts)
    
    def _get_extraction_method(self, file_type: str) -> str:
        """Get the extraction method used for the file type."""
        method_map = {
            'xlsx': 'pandas + openpyxl',
            'xls': 'pandas + openpyxl',
            'docx': 'python-docx',
            'csv': 'pandas'
        }
        return method_map.get(file_type.lower(), 'unknown')
    
    def _get_excel_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata specific to Excel files."""
        try:
            workbook = load_workbook(file_path, read_only=True)
            metadata = {
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames[:10],  # Limit to first 10
                'has_multiple_sheets': len(workbook.sheetnames) > 1
            }
            workbook.close()
            return metadata
        except Exception as e:
            self.logger.warning(f"Failed to extract Excel metadata: {str(e)}")
            return {'metadata_extraction_error': str(e)}
    
    def _get_word_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata specific to Word documents."""
        try:
            doc = Document(file_path)
            
            paragraph_count = len([p for p in doc.paragraphs if p.text.strip()])
            table_count = len(doc.tables)
            
            # Try to get document properties
            properties = {}
            try:
                core_props = doc.core_properties
                if core_props.author:
                    properties['author'] = core_props.author
                if core_props.title:
                    properties['title'] = core_props.title
                if core_props.subject:
                    properties['subject'] = core_props.subject
            except:
                pass
            
            metadata = {
                'paragraph_count': paragraph_count,
                'table_count': table_count,
                'has_tables': table_count > 0,
                'properties': properties
            }
            
            return metadata
        except Exception as e:
            self.logger.warning(f"Failed to extract Word metadata: {str(e)}")
            return {'metadata_extraction_error': str(e)}
    
    def _get_csv_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata specific to CSV files."""
        try:
            # Quick scan to get basic info without loading full file
            with open(file_path, 'r', encoding='utf-8') as f:
                first_line = f.readline()
                line_count = sum(1 for _ in f) + 1  # +1 for the first line we already read
            
            # Estimate column count from first line
            column_count = len(first_line.split(','))
            
            metadata = {
                'estimated_rows': line_count,
                'estimated_columns': column_count,
                'delimiter': ',',  # Assuming comma delimiter
                'has_header': True  # Assuming first row is header
            }
            
            return metadata
        except Exception as e:
            self.logger.warning(f"Failed to extract CSV metadata: {str(e)}")
            return {'metadata_extraction_error': str(e)}